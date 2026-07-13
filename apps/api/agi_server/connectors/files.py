from __future__ import annotations

import csv
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from agi_server.connectors.base import Health, RawRecord, SourceSchema, SyncResult

MAX_FILE_BYTES = 25 * 1024 * 1024
MAX_XLSX_EXPANDED_BYTES = 100 * 1024 * 1024
MAX_ROWS = 100_000


class ReadOnlyTabularConnector:
    """CSV/XLSX reader with bounded files and formula warnings; it exposes no write methods."""

    def __init__(
        self,
        path: Path | str,
        source_id: str,
        entity_type: str = "accounts",
        field_mapping: dict[str, str] | None = None,
    ):
        self.path = Path(path).resolve()
        self.source_id = source_id
        self.entity_type = entity_type
        self.field_mapping = field_mapping or {}

    def test_connection(self) -> Health:
        if not self.path.is_file():
            return Health(ok=False, message="Kaynak dosya bulunamadı")
        if self.path.stat().st_size > MAX_FILE_BYTES:
            return Health(ok=False, message="Kaynak dosya 25 MB sınırını aşıyor")
        if self.path.suffix.lower() not in {".csv", ".xlsx"}:
            return Health(ok=False, message="Yalnız CSV/XLSX destekleniyor")
        if self.path.suffix.lower() == ".xlsx":
            try:
                with zipfile.ZipFile(self.path) as archive:
                    if len(archive.infolist()) > 10_000:
                        return Health(ok=False, message="XLSX entry limit exceeded")
                    expanded = sum(item.file_size for item in archive.infolist())
                    if expanded > MAX_XLSX_EXPANDED_BYTES:
                        return Health(ok=False, message="XLSX expanded size limit exceeded")
            except zipfile.BadZipFile:
                return Health(ok=False, message="XLSX archive is invalid")
        return Health(ok=True, message="Read-only dosya kaynağı hazır")

    def _rows(self) -> tuple[str, list[dict[str, Any]], list[str]]:
        health = self.test_connection()
        if not health.ok:
            raise ValueError(health.message)
        warnings: list[str] = []
        if self.path.suffix.lower() == ".csv":
            with self.path.open(encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
            sheet = "csv"
        else:
            workbook = load_workbook(self.path, read_only=True, data_only=False, keep_links=False)
            sheet = workbook.sheetnames[0]
            worksheet = workbook[sheet]
            iterator = worksheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(iterator, ())]
            rows = [dict(zip(headers, values, strict=False)) for values in iterator]
        if len(rows) > MAX_ROWS:
            raise ValueError(f"Source contains more than {MAX_ROWS} rows")
        for row_index, row in enumerate(rows, start=2):
            for key, value in row.items():
                if (
                    isinstance(value, str)
                    and value.startswith(("=", "+", "-", "@"))
                    and len(warnings) < 100
                ):
                    warnings.append(f"Formula benzeri hücre: {sheet}!{key}{row_index}")
        return sheet, rows, warnings

    def discover_schema(self) -> SourceSchema:
        _, rows, _ = self._rows()
        return SourceSchema(
            source_id=self.source_id,
            entities={self.entity_type: list(rows[0]) if rows else []},
        )

    def preview(self, limit: int = 50) -> list[RawRecord]:
        return self._convert(limit=max(0, min(limit, 50)))[0]

    def preview_with_warnings(self, limit: int = 50) -> tuple[list[RawRecord], list[str]]:
        return self._convert(limit=max(0, min(limit, 50)))

    def _convert(
        self, limit: int | None = None, offset: int = 0
    ) -> tuple[list[RawRecord], list[str]]:
        sheet, rows, warnings = self._rows()
        selected = rows[offset:] if limit is None else rows[offset : offset + limit]
        mapped_rows = [
            (
                {target: row.get(source) for target, source in self.field_mapping.items()}
                if self.field_mapping
                else row
            )
            for row in selected
        ]
        records = [
            RawRecord(
                source_id=self.source_id,
                entity_type=self.entity_type,
                external_id=str(row.get("id") or row.get("ID") or row_index),
                data=row,
                locator={"kind": "tabular", "sheet": sheet, "row": row_index, "column": "*"},
            )
            for row_index, row in enumerate(mapped_rows, start=offset + 2)
        ]
        return records, warnings

    def sync(self, cursor: str | None = None) -> SyncResult:
        offset = int(cursor or 0)
        records, warnings = self._convert(offset=offset)
        return SyncResult(records=records, complete=True, warnings=warnings)

    def health(self) -> Health:
        return self.test_connection()
